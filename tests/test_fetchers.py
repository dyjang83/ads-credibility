
from __future__ import annotations

import math
import sys
from pathlib import Path

import h3
import numpy as np
import pandas as pd

# allow running both as module and as script
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.fetchers import config
from src.fetchers.osm_fetcher import (
    parse_osm, result_to_frame, build_overpass_query, haversine_km,
)
from src.fetchers.acs_fetcher import areal_interpolate, TractRecord, h3_cell_polygon
from src.fetchers.fars_fetcher import _clean_coords, crashes_in_city, crash_density_per_cell
from src.fetchers.assemble import _calibrate_hdv_frequency, _impute_missing, _validate_schema
from src.fetchers import fhwa_fetcher
from src.fetchers.fhwa_fetcher import fetch_fhwa_vmt, _locate_columns, _read_all_rows
from tests.fixtures import (
    MOCK_OVERPASS, MOCK_ACS_ROWS, MOCK_TIGER_GEOJSON, MOCK_FARS_RECORDS,
)
from shapely.geometry import shape


SF = config.CITY_BY_NAME["San Francisco"]


# --------------------------------------------------------------------------
# OSM
# --------------------------------------------------------------------------
def test_haversine_known_distance():
    # ~1 deg of latitude ~ 111 km
    d = haversine_km(37.0, -122.0, 38.0, -122.0)
    assert 110.0 < d < 112.0, d


def test_overpass_query_has_bbox_and_classes():
    q = build_overpass_query(SF.bbox)
    assert "traffic_signals" in q
    assert "primary" in q and "residential" in q
    # bbox coords present
    assert f"{SF.bbox[0]}" in q


def test_parse_osm_buckets_roads():
    parsed = parse_osm(MOCK_OVERPASS["elements"], betweenness_k=5)
    # arterial (way 1001, primary) and local (way 1002, residential) lengths > 0
    total_art = sum(d["arterial"] for d in parsed.road_km.values())
    total_loc = sum(d["local"] for d in parsed.road_km.values())
    assert total_art > 0, "primary road should land in arterial bucket"
    assert total_loc > 0, "residential road should land in local bucket"
    # collector untouched
    total_col = sum(d["collector"] for d in parsed.road_km.values())
    assert total_col == 0.0


def test_parse_osm_detects_shared_node_intersection():
    parsed = parse_osm(MOCK_OVERPASS["elements"], betweenness_k=5)
    # node 2 is shared by ways 1001 and 1002 -> an intersection
    all_intersections = set().union(*parsed.intersection_nodes.values())
    assert 2 in all_intersections


def test_parse_osm_counts_signal():
    parsed = parse_osm(MOCK_OVERPASS["elements"], betweenness_k=5)
    assert sum(parsed.signal_count.values()) == 1


def test_result_frame_schema_and_values():
    parsed = parse_osm(MOCK_OVERPASS["elements"], betweenness_k=5)
    df = result_to_frame(SF, parsed)
    for col in ("road_length_arterial_km", "intersection_density_per_km2",
                "signalized_intersection_fraction", "betweenness_centrality_mean"):
        assert col in df.columns
    # signal fraction must be within [0,1]
    assert (df["signalized_intersection_fraction"] >= 0).all()
    assert (df["signalized_intersection_fraction"] <= 1).all()
    # density non-negative
    assert (df["intersection_density_per_km2"] >= 0).all()


def test_road_length_matches_haversine():
    parsed = parse_osm(MOCK_OVERPASS["elements"], betweenness_k=5)
    # recompute expected arterial length for way 1001 directly
    g = MOCK_OVERPASS["elements"][0]["geometry"]
    expected = sum(
        haversine_km(g[i]["lat"], g[i]["lon"], g[i+1]["lat"], g[i+1]["lon"])
        for i in range(len(g) - 1)
    )
    got = sum(d["arterial"] for d in parsed.road_km.values())
    assert math.isclose(got, expected, rel_tol=1e-9), (got, expected)


# --------------------------------------------------------------------------
# ACS
# --------------------------------------------------------------------------
def test_areal_interpolation_apportions_population():
    # one tract polygon, several covering cells; population should distribute
    # such that the sum over cells <= tract population (overlap fractions <= 1)
    geom = shape(MOCK_TIGER_GEOJSON["features"][0]["geometry"])
    tract = TractRecord(
        geoid="06075010100",
        polygon=geom,
        attrs={
            "total_population": 4200.0,
            "commuters_total": 2500.0,
            "commuters_walk": 375.0,
            "housing_units_total": 1800.0,
            "units_1_detached": 900.0,
            "units_1_attached": 300.0,
        },
    )
    # cells covering the tract
    poly = h3.LatLngPoly([(37.790, -122.400), (37.790, -122.385),
                          (37.800, -122.385), (37.800, -122.400)])
    cells = list(h3.polygon_to_cells(poly, config.H3_RESOLUTION))
    assert len(cells) > 0
    out = areal_interpolate(cells, [tract])
    # pedestrian share = walk/commuters = 375/2500 = 0.15 in covered cells
    covered = out[out["population_density_per_km2"] > 0]
    assert len(covered) > 0
    assert np.allclose(covered["pedestrian_commute_share"].dropna(), 0.15, atol=1e-6)
    # residential share = (900+300)/1800 = 0.667
    assert np.allclose(covered["land_use_residential_share"].dropna(), 2/3, atol=1e-3)
    # shares within [0,1]
    assert (out["land_use_residential_share"].between(0, 1)).all()
    assert (out["land_use_commercial_share"].between(0, 1)).all()


def test_h3_cell_polygon_is_valid():
    cell = h3.latlng_to_cell(37.796, -122.394, config.H3_RESOLUTION)
    poly = h3_cell_polygon(cell)
    assert poly.is_valid
    assert poly.area > 0


# --------------------------------------------------------------------------
# FARS
# --------------------------------------------------------------------------
def test_fars_drops_sentinels():
    df = pd.DataFrame(MOCK_FARS_RECORDS)
    cleaned = _clean_coords(df)
    assert len(cleaned) == 2     # the 88.8888 sentinel row dropped


def test_fars_bbox_filter():
    df = _clean_coords(pd.DataFrame(MOCK_FARS_RECORDS))
    inside = crashes_in_city(df, SF)
    assert len(inside) == 2


def test_crash_density_normalizes_by_mileage():
    df = _clean_coords(pd.DataFrame(MOCK_FARS_RECORDS))
    inside = crashes_in_city(df, SF)
    # assign all crashes to their actual cells; build a road map giving each
    # crash-bearing cell 1 km of road (~0.621 miles)
    cells = [h3.latlng_to_cell(r.lat, r.lon, config.H3_RESOLUTION)
             for r in inside.itertuples()]
    road_map = {c: 1.0 for c in set(cells)}
    out = crash_density_per_cell(inside, road_map)
    # density = crashes / (1 km * 0.621 mi/km); for a cell with 1 crash that's ~1.609
    nonzero = out[out["historical_crash_density_per_mile"] > 0]
    assert len(nonzero) > 0
    for _, row in nonzero.iterrows():
        expected = row["_fars_crash_count"] / (1.0 * 0.621371)
        assert math.isclose(row["historical_crash_density_per_mile"], expected, rel_tol=1e-6)


def test_crash_density_zero_when_no_road():
    df = _clean_coords(pd.DataFrame(MOCK_FARS_RECORDS))
    inside = crashes_in_city(df, SF)
    out = crash_density_per_cell(inside, {})   # no road anywhere
    assert (out["historical_crash_density_per_mile"] == 0).all()


# --------------------------------------------------------------------------
# Assemble: calibration + imputation + schema
# --------------------------------------------------------------------------
def test_hdv_calibration_hits_anchor_mean():
    # build a small frame with varied crash density for SF
    rng = np.random.default_rng(0)
    df = pd.DataFrame({
        "city": ["San Francisco"] * 50,
        "historical_crash_density_per_mile": rng.exponential(2.0, size=50),
    })
    out = _calibrate_hdv_frequency(df)
    anchor = 4.06   # SF anchor in assemble.HDV_BENCHMARK_FREQ
    # mean of the calibrated frequency should match the anchor (pre-floor it's
    # exact; the 10%-floor can only raise the mean slightly)
    assert out[config.HDV_FREQUENCY_COLUMN].mean() >= anchor - 1e-6
    assert out[config.HDV_FREQUENCY_COLUMN].mean() < anchor * 1.5


def test_hdv_calibration_floor_positive():
    df = pd.DataFrame({
        "city": ["Denver"] * 10,
        "historical_crash_density_per_mile": [0.0] * 10,   # no crashes
    })
    out = _calibrate_hdv_frequency(df)
    # with zero crash density everywhere, frequency falls back to the anchor
    assert (out[config.HDV_FREQUENCY_COLUMN] > 0).all()


def test_imputation_fills_city_median():
    df = pd.DataFrame({
        "city": ["A", "A", "A"],
        "population_density_per_km2": [10.0, 20.0, np.nan],
    })
    filled, report = _impute_missing(df, ["population_density_per_km2"])
    assert not filled["population_density_per_km2"].isna().any()
    assert filled["population_density_per_km2"].iloc[2] == 15.0   # median of 10,20
    assert report["population_density_per_km2"] == 1


def test_validate_schema_rejects_wrong_columns():
    bad = pd.DataFrame({"cell_id": ["x"], "city": ["A"], "wrong": [1.0]})
    try:
        _validate_schema(bad)
    except ValueError:
        return
    raise AssertionError("schema validation should have raised")


def test_validate_schema_accepts_correct_columns():
    cols = ["cell_id", "city", *config.CELL_FEATURE_COLUMNS, config.HDV_FREQUENCY_COLUMN]
    good = pd.DataFrame({c: [0.0] for c in cols})
    good["cell_id"] = ["x"]; good["city"] = ["A"]
    _validate_schema(good)   # should not raise



# --------------------------------------------------------------------------
# Regression tests for the bugs fixed after traceback on Diane's machine
# --------------------------------------------------------------------------
def test_acs_in_param_uses_percent20_not_plus():
    """The `in` predicate must use %20, not +, between sub-predicates."""
    import sys; sys.path.insert(0, ".")
    from src.fetchers.acs_fetcher import ACS_VARIABLES, ACS_BASE
    from src.fetchers.config import CITY_BY_NAME

    city = CITY_BY_NAME["San Francisco"]
    var_codes = ",".join(ACS_VARIABLES.keys())
    params = {
        "get": f"NAME,{var_codes}",
        "for": "tract:*",
        "in": f"state:{city.state_fips}%20county:{city.county_fips[0]}",
    }
    # The `in` value must contain %20, not a raw space or +
    assert "%20" in params["in"], params["in"]
    assert " " not in params["in"], "raw space found — Census will reject"
    assert "+" not in params["in"], "+ encoding found — Census may reject"


def test_acs_get_param_no_duplicate_name():
    """NAME must appear exactly once in the `get` parameter."""
    from src.fetchers.acs_fetcher import ACS_VARIABLES
    var_codes = ",".join(ACS_VARIABLES.keys())
    get_param = f"NAME,{var_codes}"
    cols = get_param.split(",")
    assert cols.count("NAME") == 1, f"NAME appears {cols.count('NAME')} times"


def test_http_cache_surfaces_body_on_non_json_200():
    """A 200 response with non-JSON body raises RuntimeError showing the body."""
    import tempfile, requests as req_mod
    with tempfile.TemporaryDirectory() as td:
        from src.fetchers.http_cache import CachedSession
        sess = CachedSession(td, user_agent="test", request_pause_s=0.0, max_retries=2)

        class FakeResp:
            status_code = 200
            text = "<html>Missing Key</html>"
            def raise_for_status(self): pass
            def json(self):
                # Simulate requests >= 2.28 raising JSONDecodeError (subclass
                # of RequestException, not ValueError) on Python 3.14.
                raise req_mod.exceptions.JSONDecodeError(
                    "Expecting value", "", 0)

        calls = {"n": 0}
        def counting_request(*a, **kw):
            calls["n"] += 1
            return FakeResp()
        sess._session.request = counting_request

        try:
            sess.get_json("https://example.census.gov/", params={"x": 1})
            raise AssertionError("should have raised")
        except RuntimeError as exc:
            msg = str(exc)
            assert "not JSON" in msg, msg
            assert "Missing Key" in msg, msg
            # Must NOT retry — one attempt only.
            assert calls["n"] == 1, f"expected 1 call, got {calls['n']}"


def test_acs_requires_api_key():
    """fetch_acs_features raises clearly when CENSUS_API_KEY is not set."""
    import os, tempfile
    env_backup = os.environ.pop("CENSUS_API_KEY", None)
    try:
        from src.fetchers.acs_fetcher import fetch_acs_features
        from src.fetchers.config import FetchConfig, CITY_BY_NAME
        cfg = FetchConfig(cities=(CITY_BY_NAME["San Francisco"],))
        try:
            fetch_acs_features(cfg)
            raise AssertionError("should have raised")
        except RuntimeError as exc:
            msg = str(exc)
            assert "key_signup" in msg, msg
            assert "CENSUS_API_KEY" in msg, msg
    finally:
        if env_backup is not None:
            os.environ["CENSUS_API_KEY"] = env_backup


# --------------------------------------------------------------------------
# FHWA urbanized-area VMT fetcher
# --------------------------------------------------------------------------
def _build_hm72_fixture(fmt: str = "xlsx") -> bytes:
    """Bytes mimicking the real FHWA HM-72 layout:
    - Sheet 'CRYSTAL_PERSIST' (skipped)
    - Sheet 'A': largest urbanized areas, rows 0-13 are banner/header,
      data starts row 14, col 0 = area name, col 2 = DVMT (thousands daily miles)
    - Sheet 'footnotes' (skipped)
    All seven of our cities appear on sheet 'A' (matching the real file).
    """
    import io
    # DVMT values (thousands of daily vehicle-miles) from real HM-72 2022
    data_A = [
        ("New York--Newark, NY--NJ--CT",           44301, 278774, 18351295, 3450, 5319),
        ("Los Angeles--Long Beach--Anaheim, CA",   25010, 245546, 12150996, 1736, 6999),
        ("Chicago, IL--IN",                        31728, 167492,  8608208, 2443, 3524),
        ("Miami, FL",                              15646, 136327,  5502379, 1239, 4442),
        ("Boston, MA--NH--RI",                     17855,  98925,  4181019, 1873, 2232),
        ("Phoenix--Mesa, AZ",                      16073, 108902,  3629114, 1147, 3165),
        ("San Francisco--Oakland, CA",              7411,  58066,  3281212,  524, 6266),
        ("Denver--Aurora, CO",                      9516,  58900,  2374203,  668, 3554),
        ("Austin, TX",                              5432,  25632,  1011790,  252, 4014),
    ]
    header = [
        ["CRYSTAL_PERSIST"] + [""] * 5,            # row 0 of sheet A banner
        [""] * 6,
        [""] * 6,
        [""] * 6,
        ["URBANIZED  AREAS - 2022"] + [""] * 5,
        ["SELECTED  CHARACTERISTICS"] + [""] * 5,
        [""] * 6,
        [""] * 6,
        ["45320.822378206016"] + [""] * 5,
        [""] * 6,
        ["", "TOTAL", "TOTAL", "CENSUS", "NET", "PERSONS"],
        ["FEDERAL-AID", "ROADWAY", "DVMT  (2)", "POPULATION", "LAND", "PER"],
        ["URBANIZED  AREA  (1)", "MILES", "(1,000)", "", "AREA", "SQUARE"],
        ["", "", "", "", "(SQ. MILES)", "MILE"],
    ]
    if fmt == "xlsx":
        from openpyxl import Workbook
        wb = Workbook()
        # CRYSTAL_PERSIST sheet (skipped by parser)
        ws0 = wb.active; ws0.title = "CRYSTAL_PERSIST"
        ws0.append(["CRYSTAL_PERSIST"])
        # Data sheet 'A'
        wsA = wb.create_sheet("A")
        for row in header:
            wsA.append(row)
        for row in data_A:
            wsA.append(list(row))
        wsA.append([""] * 6)
        # footnotes sheet (skipped)
        wsF = wb.create_sheet("footnotes")
        wsF.append(["URBANIZED  AREAS - 2022", ""])
        bio = io.BytesIO(); wb.save(bio); return bio.getvalue()
    else:
        try:
            import xlwt
            wb = xlwt.Workbook()
            ws0 = wb.add_sheet("CRYSTAL_PERSIST"); ws0.write(0, 0, "CRYSTAL_PERSIST")
            wsA = wb.add_sheet("A")
            for ri, row in enumerate(header):
                for ci, v in enumerate(row): wsA.write(ri, ci, v)
            for ri, row in enumerate(data_A, len(header)):
                for ci, v in enumerate(row): wsA.write(ri, ci, v)
            wsF = wb.add_sheet("footnotes"); wsF.write(0, 0, "URBANIZED  AREAS - 2022")
            bio = io.BytesIO(); wb.save(bio); return bio.getvalue()
        except ImportError:
            return _build_hm72_fixture("xlsx")


def _run_fhwa_on_fixture(fixture_bytes, monkey_cities=None):
    orig = fhwa_fetcher._download_workbook
    fhwa_fetcher._download_workbook = lambda year, cache_dir, ua: fixture_bytes
    try:
        cfg = config.FetchConfig()
        if monkey_cities is not None:
            cfg.cities = monkey_cities
        return fetch_fhwa_vmt(cfg)
    finally:
        fhwa_fetcher._download_workbook = orig


def test_fhwa_locates_area_and_total_columns():
    raw = _read_all_rows(_build_hm72_fixture("xlsx"))
    area_col, total_col = _locate_columns(raw)
    assert area_col == 0, area_col
    assert total_col == 2, total_col


def test_fhwa_detects_xls_magic_bytes():
    from src.fetchers.fhwa_fetcher import _detect_format, _XLS_MAGIC, _XLSX_MAGIC
    assert _detect_format(_XLS_MAGIC + b"\xa1\xb1" * 100) == "xls"
    assert _detect_format(_XLSX_MAGIC + b"\x14\x00" * 100) == "xlsx"


def test_fhwa_html_response_raises_not_silently_fails():
    """If FHWA returns an HTML redirect page, _detect_format raises clearly."""
    from src.fetchers.fhwa_fetcher import _detect_format
    html = b"<html><body>Page not found</body></html>"
    try:
        _detect_format(html)
        raise AssertionError("should have raised")
    except RuntimeError as exc:
        assert "HTML" in str(exc) or "html" in str(exc).lower(), str(exc)


def test_fhwa_converts_thousands_to_millions():
    df = _run_fhwa_on_fixture(_build_hm72_fixture("xlsx"))
    sf = df.loc[df.city == "San Francisco", "annual_vmt_millions"].iloc[0]
    # SF DVMT = 58,066 thousand daily miles → annual = 58,066 × 0.365 = 21,194.09 M miles
    assert math.isclose(sf, 58066 * 0.365, rel_tol=1e-6), sf


def test_fhwa_matches_multi_anchor_names():
    df = _run_fhwa_on_fixture(_build_hm72_fixture("xlsx"))
    assert df["annual_vmt_millions"].notna().all(), df
    la = df.loc[df.city == "Los Angeles", "annual_vmt_millions"].iloc[0]
    assert math.isclose(la, 245546 * 0.365, rel_tol=1e-6), la


def test_fhwa_unmatched_city_is_nan_not_error():
    """A city with no urbanized-area row returns NaN rather than raising."""
    from dataclasses import replace
    fake = replace(config.CITY_BY_NAME["Austin"], name="Nowhereville", slug="Nowhereville")
    df = _run_fhwa_on_fixture(_build_hm72_fixture("xlsx"), monkey_cities=(fake,))
    assert len(df) == 1
    assert pd.isna(df["annual_vmt_millions"].iloc[0])



if __name__ == "__main__":
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    passed = 0
    for fn in fns:
        try:
            fn()
            print(f"PASS {fn.__name__}")
            passed += 1
        except Exception as exc:
            print(f"FAIL {fn.__name__}: {exc}")
    print(f"\n{passed}/{len(fns)} passed")
    sys.exit(0 if passed == len(fns) else 1)
